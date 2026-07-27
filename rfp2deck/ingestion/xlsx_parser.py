from __future__ import annotations

import re
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Union

from openpyxl import load_workbook

from rfp2deck.core.logging import get_logger
from rfp2deck.core.schemas import ClarificationRecord

log = get_logger(__name__)

_QUESTION_HEADERS = {
    "question",
    "questions",
    "vendor question",
    "vendor query",
    "query",
    "clarification",
    "clarification question",
    "clarification requested",
}
_ANSWER_HEADERS = {
    "answer",
    "response",
    "customer answer",
    "customer response",
    "client answer",
    "client response",
    "reply",
}
_ID_HEADERS = {
    "id",
    "no",
    "number",
    "serial no",
    "s no",
    "question id",
    "query id",
    "clarification id",
}


@dataclass
class ParsedWorkbook:
    text: str
    sheet_count: int
    row_count: int
    clarifications: List[ClarificationRecord] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


def _normalise_header(value: object) -> str:
    text = re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().lower())
    return re.sub(r"\s+", " ", text).strip()


def _is_question_header(value: str) -> bool:
    if value in _QUESTION_HEADERS:
        return True
    has_question_term = any(term in value.split() for term in ("question", "questions", "query"))
    has_answer_term = any(term in value.split() for term in ("answer", "response", "reply"))
    return has_question_term and not has_answer_term


def _is_answer_header(value: str) -> bool:
    if value in _ANSWER_HEADERS:
        return True
    words = set(value.split())
    has_answer_term = bool(words.intersection({"answer", "response", "reply"}))
    looks_like_deadline = bool(words.intersection({"date", "deadline", "due"}))
    return has_answer_term and not looks_like_deadline


def _display_value(value: object) -> str:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except (TypeError, ValueError):
            pass
    return re.sub(r"\s+", " ", str(value)).strip()


def _find_header_row(rows: List[List[str]]) -> tuple[int, Dict[str, int]] | None:
    for row_index, values in enumerate(rows[:25]):
        normalised = [_normalise_header(value) for value in values]
        question_col = next(
            (index for index, value in enumerate(normalised) if _is_question_header(value)),
            None,
        )
        answer_col = next(
            (index for index, value in enumerate(normalised) if _is_answer_header(value)),
            None,
        )
        if question_col is None or answer_col is None:
            continue
        id_col = next(
            (index for index, value in enumerate(normalised) if value in _ID_HEADERS),
            None,
        )
        columns = {"question": question_col, "answer": answer_col}
        if id_col is not None:
            columns["id"] = id_col
        return row_index, columns
    return None


def parse_xlsx(
    path_or_bytes: Union[Path, bytes],
    *,
    document_id: str = "xlsx",
) -> ParsedWorkbook:
    source = "bytes" if isinstance(path_or_bytes, bytes) else str(path_or_bytes)
    try:
        if isinstance(path_or_bytes, bytes):
            workbook = load_workbook(BytesIO(path_or_bytes), data_only=True, read_only=True)
        else:
            workbook = load_workbook(path_or_bytes, data_only=True, read_only=True)
    except Exception:
        log.exception("Failed to open XLSX (source=%s)", source)
        raise

    text_lines: List[str] = []
    clarifications: List[ClarificationRecord] = []
    warnings: List[str] = []
    total_rows = 0

    for worksheet in workbook.worksheets:
        sheet_name = worksheet.title
        rows: List[List[str]] = []
        row_numbers: List[int] = []
        for row_number, cells in enumerate(worksheet.iter_rows(values_only=True), start=1):
            values = [_display_value(value) for value in cells]
            while values and not values[-1]:
                values.pop()
            if not any(values):
                continue
            rows.append(values)
            row_numbers.append(row_number)

        total_rows += len(rows)
        text_lines.append(f'\n--- SHEET "{sheet_name}" ---')
        if not rows:
            text_lines.append("[EMPTY SHEET]")
            continue

        header_info = _find_header_row(rows)
        headers: List[str] = []
        if header_info is not None:
            header_index, columns = header_info
            headers = rows[header_index]
            for relative_index in range(header_index + 1, len(rows)):
                values = rows[relative_index]
                actual_row = row_numbers[relative_index]
                question = values[columns["question"]] if columns["question"] < len(values) else ""
                answer = values[columns["answer"]] if columns["answer"] < len(values) else ""
                if not question and not answer:
                    continue
                raw_id = ""
                if "id" in columns and columns["id"] < len(values):
                    raw_id = values[columns["id"]]
                clarification_id = raw_id or f"{document_id}-{sheet_name}-row-{actual_row}"
                clarifications.append(
                    ClarificationRecord(
                        clarification_id=str(clarification_id),
                        document_id=document_id,
                        question=question,
                        customer_response=answer,
                        source_ref=f'{document_id} / sheet "{sheet_name}" / row {actual_row}',
                        status="active" if answer else "unresolved",
                    )
                )

        for relative_index, values in enumerate(rows):
            actual_row = row_numbers[relative_index]
            labelled_values: List[str] = []
            for column_index, value in enumerate(values):
                if not value:
                    continue
                header = ""
                if headers and relative_index > (header_info[0] if header_info else -1):
                    if column_index < len(headers):
                        header = headers[column_index]
                label = header or f"Column {column_index + 1}"
                labelled_values.append(f"{label}: {value}")
            text_lines.append(
                f'[SHEET "{sheet_name}"][ROW {actual_row}] ' + " | ".join(labelled_values)
            )

    sheet_count = len(workbook.sheetnames)
    workbook.close()
    if not clarifications:
        warnings.append(
            "No question/customer-response column pair was detected; workbook rows remain available as evidence."
        )
    parsed = ParsedWorkbook(
        text="\n".join(text_lines).strip(),
        sheet_count=sheet_count,
        row_count=total_rows,
        clarifications=clarifications,
        warnings=warnings,
    )
    log.info(
        "Parsed XLSX (source=%s): %d sheets, %d non-empty rows, %d clarification records",
        source,
        parsed.sheet_count,
        parsed.row_count,
        len(parsed.clarifications),
    )
    return parsed
